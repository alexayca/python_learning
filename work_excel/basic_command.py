
# importamos la libreria https://openpyxl.readthedocs.io/
# as excel_lib es opcional
import openpyxl as excel_lib


# ----- / ----- / ----- / ----- / ----- / ----- / ----- / ----- / -----   
# Para abrir el archivo Excel es load_workbook().
excel_document = excel_lib.load_workbook('sample.xlsx')

# el tipo de dato devuelto por el metodo
print(type(excel_document))

# obtener los nombres de las pestanas (hojas) del libro abierto
nombres_pestanias = excel_document.sheetnames
print(type(nombres_pestanias) , nombres_pestanias)


# ----- / ----- / ----- / ----- / ----- / ----- / ----- / ----- / -----  
# ----- / ----- / ----- / ----- / ----- / ----- / ----- / ----- / -----   
# CREAR UN ARCHIVO DE EXCEL con una instancia de un libro en blanco
#  sin escribirlo en disco
wb = excel_lib.Workbook()

# El libro recién creado contiene una única hoja que está activa.
# Para acceder al nombre de la hoja activa usamos el atributo active.
hoja = wb.active
print(f'Hoja activa: {hoja.title}')

# para modificar el nombre de la hoja activa
hoja.title = "Valores"
print(f'Hoja activa: {wb.active.title}')


# ----- / ----- / ----- / ----- / ----- / ----- / ----- / ----- / -----   
# CREAR UNA HOJA extra en el libro activo workbook
# Añade la hoja 'Hoja' al final (por defecto)
hoja1 = wb.create_sheet("Hoja")

# Añade la hoja 'Hoja' en la primera posición. Como el nombre
# 'Hoja' ya existe, le añade el número 1 al final del nombre
# Sera la hoja activa
hoja2 = wb.create_sheet("Hoja", 0)

# Añade la hoja 'Otra hoja' en la posición 1
wb.create_sheet(index=1, title="Otra hoja")

# COPIAR UNA HOJA
origen = wb.active
nueva_hoja = wb.copy_worksheet(origen)

# Muestra los nombres de las hojas
# get_sheet_names() is deprecated use sheetnames
print(wb.sheetnames)


# ----- / ----- / ----- / ----- / ----- / ----- / ----- / ----- / -----   
# ACCEDER A UNA HOJA
# asigna y se muestra la hoja activa
hoja = wb.active  # Es la hoja que está en el índice 0
print(f'Hoja activa: {hoja.title}')

# Cambiamos a 'Otra hoja' como hoja activa
hoja = wb['Otra hoja']
wb.active = hoja    # asignamos la hoja activa
print(f'Hoja activa: {wb.active.title}')

# Podemos iterar sobre todas las hojas al ser una lista
print(wb.sheetnames)    # mostramos la lista de las hojas del libro wb

for nombres_hoja in wb:
    print(nombres_hoja.title)


# ----- / ----- / ----- / ----- / ----- / ----- / ----- / ----- / -----   
# ACCEDER A UNA CELDA
# IMPORTANTE: Cuando se crea un libro, este no contiene ninguna celda. 
# Las celdas se van creando en memoria a medida que se va accediendo a ellas, 
# aunque no contengan ningún valor.

# Es posible acceder a una celda tratando a la hoja como un diccionario, 
# La clave es la celda: letra de la columna y el número de la fila.
hoja = wb.active
a1 = hoja["A1"]
print(a1.value) # al no existir valor: none

# También es posible acceder a una celda usando la notación fila, columna 
# con el método cell()
b2 = hoja.cell(row=2, column=2)
print(b2.value)


# ----- / ----- / ----- / ----- / ----- / ----- / ----- / ----- / -----   
# ESCRIBIR VALORES EN UNA CELDA
# Se puede hacer de tres formas diferentes

# 1.- Asignando el valor directamente a la celda
hoja["A1"] = 10 # asigna 10 a la celda A1
a1 = hoja["A1"] # asigna el valor de la celda A1 a la var a1
print(a1.value) # para mostrarla con la funcion print

# 2.- Usando la notación fila, columna con el argumento value
b1 = hoja.cell(row=1, column=2, value=20)   # B1
print(b1.value)

# 3.- Actualizando la propiedad value de una celda
c1 = hoja.cell(row=1, column=3) # se coloca en la celda deseada C1
c1.value = 30           # asigna el valor
print(c1.value)


# ----- / ----- / ----- / ----- / ----- / ----- / ----- / ----- / -----   
# GUARDAR UNA LISTA DE VALORES
# en Python suele ocurrir que los datos están almacenados en listas o tuplas.
productos = [
    ('producto_1', 'a859', 1500, 9.95),
    ('producto_2', 'b125', 600, 4.95),
    ('producto_3', 'c764', 200, 19.95),
    ('producto_4', 'd399', 2000, 49.95)
]
# Podemos usar el método append() de un objeto hoja.
wb2 = excel_lib.Workbook()  # creamos el libro
hoja = wb2.active
# Crea la fila del encabezado con los títulos
hoja.append(('Nombre', 'Referencia', 'Stock', 'Precio'))
for producto in productos:
    # producto es una tupla con los valores de un producto 
    hoja.append(producto)


# ----- / ----- / ----- / ----- / ----- / ----- / ----- / ----- / -----   
# GUARDAR UN LIBRO DE EXCEL
# Llamar al método save() del workbook indicando el nombre del archivo. 
# Esto guardará el libro con todas las hojas y los datos de cada una de ellas.
# si existe el archivo lo sobreescribe
wb.save('tutorial.xlsx')
wb2.save('productos.xlsx')

